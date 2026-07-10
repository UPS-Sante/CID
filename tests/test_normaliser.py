# -*- coding: utf-8 -*-
"""Tests du normaliseur. Execution : python tests/test_normaliser.py
Construit des fichiers de test dans un dossier temporaire, lance la
normalisation et verifie les enregistrements produits et les anomalies
attendues. Aucune dependance au-dela de celles de requirements.txt.
"""

import csv
import json
import os
import shutil
import subprocess
import sys
import tempfile

import openpyxl

RACINE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

PROVINCES = ["BATHA", "BARH-EL-GAZAL", "BORKOU", "CHARI BAGUIRMI", "ENNEDI EST",
             "ENNEDI OUEST", "GUERA", "HADJER LAMIS", "KANEM", "LAC",
             "LOGONE OCCIDENTAL", "LOGONE ORIENTAL", "MANDOUL", "MAYO KEBBI EST",
             "MAYO KEBBI OUEST", "MOYEN CHARI", "N'DJAMENA", "OUADDAI", "SALAMAT",
             "SILA", "TANDJILE", "TIBESTI", "WADI FIRA", "TOTAL"]


def construire_jeu_de_test(brut):
    """Fichiers couvrant les cas sensibles : zeros, virgules, nd, feuille de
    notes, province inconnue, deux feuilles matrices dans un classeur."""
    os.makedirs(os.path.join(brut, "DSR"))
    os.makedirs(os.path.join(brut, "NUTRITION"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sante maternelle"
    ws.cell(1, 1, "Provinces")
    ws.cell(1, 2, "Deces maternels notifies")
    ws.cell(1, 3, "Couverture CPN4")
    for i, p in enumerate(PROVINCES, start=2):
        ws.cell(i, 1, p)
        ws.cell(i, 2, 0 if i == 2 else i)          # un zero explicite
        ws.cell(i, 3, "84,5" if i == 3 else "nd" if i == 4 else 50.0)
    ws2 = wb.create_sheet("Notes")
    ws2.cell(1, 1, "Observations du mois")
    ws2.cell(2, 1, "RAS")
    wb.save(os.path.join(brut, "DSR", "Indicateurs SRMNIA_Juin 2026.xlsx"))

    with open(os.path.join(brut, "NUTRITION", "Nutrition 2026-06.csv"),
              "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["periode", "province", "indicateur", "valeur"])
        for p in PROVINCES[:5]:
            w.writerow(["2026-06", p, "MAS admise", "12,5"])
        w.writerow(["2026-06", "ZONE X", "MAS admise", "9"])   # anomalie attendue


def principal():
    dossier = tempfile.mkdtemp(prefix="test_cid_")
    brut = os.path.join(dossier, "brut")
    data = os.path.join(dossier, "data")
    construire_jeu_de_test(brut)

    config = {"DSR": "matrice_provinces", "NUTRITION": "format_long"}
    with open(os.path.join(dossier, "config_sources.json"), "w", encoding="utf-8") as f:
        json.dump(config, f)
    shutil.copy(os.path.join(RACINE, "referentiel.json"), dossier)

    r = subprocess.run([sys.executable, os.path.join(RACINE, "scripts", "normaliser.py")],
                       cwd=dossier, capture_output=True, text=True,
                       env={**os.environ, "SOURCE_DIR": brut, "DATA_DIR": data})
    assert r.returncode == 0, r.stderr

    base = json.load(open(os.path.join(data, "consolide.json"), encoding="utf-8"))["enregistrements"]
    anomalies = json.load(open(os.path.join(data, "rapport_validation.json"), encoding="utf-8"))["anomalies"]

    assert any(e["valeur"] == 0 and e["province"] == "Batha" for e in base), "zero perdu"
    assert any(e["valeur"] == 84.5 for e in base), "virgule decimale non convertie"
    assert any(e["province"] == "Barh El Gazal" for e in base), "variante de province non rattachee"
    assert any(e["province"] == "National" for e in base), "ligne TOTAL non rattachee a National"
    assert any(e["source"] == "NUTRITION" and e["valeur"] == 12.5 for e in base), "lecteur format long"
    types = [a["type"] for a in anomalies]
    assert "province_inconnue" in types, "province inconnue non signalee"
    assert "matrice_non_reconnue" in types, "feuille de notes non signalee"
    assert not any(e["valeur"] is None for e in base), "valeur nulle dans la base"

    shutil.rmtree(dossier)
    print(f"TESTS NORMALISEUR : OK ({len(base)} enregistrements, {len(anomalies)} anomalies attendues)")


if __name__ == "__main__":
    principal()
