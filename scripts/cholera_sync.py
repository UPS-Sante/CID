# -*- coding: utf-8 -*-
"""Recalcule les jeux de donnees du cholera a partir de la liste lineaire
deposee dans le dossier Drive de surveillance.

Le fichier attendu est un classeur dont une feuille porte la liste lineaire
des cas, une ligne par patient. Le script normalise les libelles, agrege par
jour, semaine epidemiologique et mois, puis ecrit les fichiers consommes par
le tableau de bord.

Sorties : cholera_data.json, chol_periodes.json, chol_series.json

Environnement :
  BRUT  repertoire des fichiers telecharges depuis Drive (defaut : brut)
"""

import collections
import json
import os
import re
import sys
import unicodedata
from datetime import date, datetime, timezone

import openpyxl

BRUT = os.environ.get("BRUT", "brut")
# La liste lineaire est deposee par le COUSP dans le sous-dossier du meme nom.
DOSSIER_COUSP = os.environ.get("DOSSIER_COUSP", "COUSP")
MOTIFS = ("cholera", "choléra", "linear", "lineaire", "linelist")

# Colonnes de la liste lineaire, reperees par position (1-indexe)
COL = {
    "semaine": 2, "nom": 3, "age_annees": 4, "age_mois": 5, "tranche": 6,
    "sexe": 7, "province": 8, "district": 9, "fosa": 10, "village": 11,
    "date_symptomes": 15, "date_consultation": 16,
    "diarrhee": 19, "vomissements": 20, "crampes": 21, "fievre": 22,
    "asthenie": 23, "soif": 24, "deshydratation": 26, "hospitalisation": 27,
    "evolution": 34,
}

# Variantes de libelles a fusionner, validees manuellement
FUSION_VILLAGES = {"Djougouli-Adre": "Djougoulié-Adré", "Djougoulié": "Djougoulié-Adré",
                   "Farcha Ethiopie": "Farcha"}


def trouver_classeur():
    """Retient le classeur le plus recent depose par le COUSP.

    La recherche porte d'abord sur le sous-dossier COUSP, ou le centre depose
    la liste lineaire. Si ce dossier n'existe pas encore dans l'arborescence
    synchronisee, la recherche s'elargit a l'ensemble des fichiers telecharges.
    """
    def collecter(base):
        trouves = []
        for racine, _, fichiers in os.walk(base):
            for f in fichiers:
                if not f.lower().endswith((".xlsx", ".xls")):
                    continue
                if any(m in f.lower() for m in MOTIFS):
                    chemin = os.path.join(racine, f)
                    trouves.append((os.path.getmtime(chemin), chemin))
        return trouves

    cousp = None
    for racine, dossiers, _ in os.walk(BRUT):
        for d in dossiers:
            if d.upper() == DOSSIER_COUSP.upper():
                cousp = os.path.join(racine, d)
                break
        if cousp:
            break

    candidats = collecter(cousp) if cousp else []
    if not candidats:
        if cousp:
            print(f"Aucun classeur cholera dans {cousp}, recherche elargie.")
        candidats = collecter(BRUT)
    if not candidats:
        return None
    return sorted(candidats)[-1][1]


def feuille_liste_lineaire(classeur):
    """Retient la feuille comportant le plus de lignes renseignees."""
    meilleure, taille = None, 0
    for nom in classeur.sheetnames:
        ws = classeur[nom]
        if ws.max_row > taille and ws.max_column >= 30:
            meilleure, taille = ws, ws.max_row
    return meilleure


def cle_village(v):
    x = unicodedata.normalize("NFD", str(v)).encode("ascii", "ignore").decode().upper()
    return re.sub(r"[^A-Z0-9]", "", x)


def normaliser_villages(lignes):
    """Fusionne les variantes de casse et d'orthographe d'une meme localite."""
    groupes = collections.defaultdict(list)
    for r in lignes:
        if r["village"]:
            groupes[cle_village(r["village"])].append(r["village"])
    canon = {}
    for _, variantes in groupes.items():
        frequences = collections.Counter(variantes).most_common()
        forme = next((x for x, _ in frequences
                      if not x.isupper() and not x.islower()), frequences[0][0])
        for v in set(variantes):
            canon[v] = forme
    for r in lignes:
        if r["village"]:
            v = canon.get(r["village"], r["village"])
            r["village"] = FUSION_VILLAGES.get(v, v)


def valeur_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def oui_non(v):
    if v is None:
        return None
    s = str(v).strip().lower()
    return 1 if s.startswith("oui") else 0 if s.startswith("non") else None


def nombre(v):
    try:
        return float(str(v).strip().replace(",", "."))
    except (TypeError, ValueError):
        return None


def province_normalisee(v):
    if not v:
        return None
    s = unicodedata.normalize("NFD", str(v)).encode("ascii", "ignore").decode().upper()
    if "HADJER" in s:
        return "Hadjer Lamis"
    if "DJAMENA" in s:
        return "N'Djamena"
    return str(v).strip()


def lire(ws):
    lignes = []
    for r in range(2, ws.max_row + 1):
        if not ws.cell(r, COL["nom"]).value:
            continue
        cel = lambda c: ws.cell(r, COL[c]).value
        evolution = str(cel("evolution")).strip() if cel("evolution") is not None else None
        ds, dc = valeur_date(cel("date_symptomes")), valeur_date(cel("date_consultation"))
        lignes.append({
            "semaine": nombre(cel("semaine")),
            "age": nombre(cel("age_annees")),
            "tranche": str(cel("tranche")).replace("\xa0", " ").strip() if cel("tranche") else None,
            "sexe": str(cel("sexe")).strip()[:1].upper() if cel("sexe") else None,
            "province": province_normalisee(cel("province")),
            "district": str(cel("district")).strip() if cel("district") else None,
            "village": str(cel("village")).strip() if cel("village") else None,
            "date": dc or ds,
            "deshydratation": nombre(cel("deshydratation")),
            "hospitalisation": oui_non(cel("hospitalisation")),
            "evolution": evolution,
            "deces": 1 if evolution == "3" else (0 if evolution in ("1", "2", "4") else None),
        })
    return lignes


def agreger(sous):
    n = len(sous)
    issue = [r for r in sous if r["deces"] is not None]
    deces = sum(r["deces"] for r in issue)
    desh = collections.Counter(r["deshydratation"] for r in sous if r["deshydratation"])
    evol = collections.Counter(r["evolution"] for r in sous if r["evolution"])
    districts = collections.Counter(r["district"] for r in sous if r["district"])
    provinces = collections.Counter(r["province"] for r in sous if r["province"])
    ages = sorted(r["age"] for r in sous if r["age"] is not None)
    sexe = collections.Counter(r["sexe"] for r in sous if r["sexe"])
    return {
        "cas": n, "deces": deces, "cas_issue": len(issue),
        "letalite": round(deces / len(issue) * 100, 1) if issue else None,
        "gueris": evol.get("1", 0), "malades": evol.get("2", 0),
        "desh": {"legere": desh.get(1.0, 0), "moderee": desh.get(2.0, 0),
                 "severe": desh.get(3.0, 0)},
        "districts": [{"nom": k, "cas": v,
                       "deces": sum(1 for r in sous if r["district"] == k and r["deces"] == 1)}
                      for k, v in districts.most_common()],
        "provinces": [{"nom": k, "cas": v,
                       "deces": sum(1 for r in sous if r["province"] == k and r["deces"] == 1)}
                      for k, v in provinces.most_common()],
        "age_median": ages[len(ages) // 2] if ages else None,
        "age_min": min(ages) if ages else None,
        "age_max": max(ages) if ages else None,
        "sexe": {"M": sexe.get("M", 0), "F": sexe.get("F", 0)},
        "ratio_fh": round(sexe.get("F", 0) / sexe.get("M", 1), 2) if sexe.get("M") else None,
        "villages": len(set(r["village"] for r in sous if r["village"])),
    }


def ecrire_periodes(lignes, reference):
    valides = [r for r in lignes if r["date"]]
    if not valides:
        raise SystemExit("Aucune ligne datee : verifier les colonnes de dates.")
    groupes = {"jour": {}, "semaine": {}, "mois": {}}
    for r in valides:
        d = r["date"]
        iso = d.isocalendar()
        for typ, cle in (("jour", d.isoformat()),
                         ("semaine", f"{iso[0]}-S{iso[1]:02d}"),
                         ("mois", f"{d.year}-{d.month:02d}")):
            groupes[typ].setdefault(cle, []).append(r)

    sortie = {"meta": {
        "debut": min(r["date"] for r in valides).isoformat(),
        "fin": max(r["date"] for r in valides).isoformat(),
        "n": len(valides), "source": "Liste lineaire COUSP",
        "actualise": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
    }}
    for typ, contenu in groupes.items():
        sortie[typ] = {}
        for cle in sorted(contenu):
            cumul = [r for r in valides if (
                r["date"].isoformat() <= cle if typ == "jour"
                else f"{r['date'].isocalendar()[0]}-S{r['date'].isocalendar()[1]:02d}" <= cle
                if typ == "semaine"
                else f"{r['date'].year}-{r['date'].month:02d}" <= cle)]
            sortie[typ][cle] = {"periode": agreger(contenu[cle]), "cumul": agreger(cumul)}
    sortie["global"] = agreger(valides)
    sortie["cles"] = {t: sorted(groupes[t]) for t in groupes}
    sortie["sitrep"] = reference
    with open("chol_periodes.json", "w", encoding="utf-8") as f:
        json.dump(sortie, f, ensure_ascii=False)
    return sortie


def ecrire_riposte(lignes, periodes):
    """Fichier consomme par les vues de riposte et de foyers."""
    g = periodes["global"]
    sortie = {
        "meta": {"maladie": "Choléra",
                 "periode": f"Du {periodes['meta']['debut']} au {periodes['meta']['fin']}",
                 "zone": " et ".join(p["nom"] for p in g["provinces"]),
                 "source": "Liste lineaire COUSP", "seuil_letalite": 1.0,
                 "actualise": periodes["meta"]["actualise"]},
        "global": {"cas": g["cas"], "deces": g["deces"],
                   "cas_issue_connue": g["cas_issue"], "sans_issue": g["cas"] - g["cas_issue"],
                   "gueris": g["gueris"], "malades": g["malades"],
                   "létalité": g["letalite"], "villages": g["villages"],
                   "hospitalises": sum(1 for r in lignes if r["hospitalisation"] == 1)},
        "provinces": [dict(p, létalité=round(p["deces"] / p["cas"] * 100, 1) if p["cas"] else 0)
                      for p in g["provinces"]],
        "districts": [dict(d, létalité=round(d["deces"] / d["cas"] * 100, 1) if d["cas"] else 0)
                      for d in g["districts"]],
        "hebdo": [{"sem": int(c.split("-S")[1]), "cas": periodes["semaine"][c]["periode"]["cas"]}
                  for c in periodes["cles"]["semaine"]],
        "age": [{"tranche": t, "cas": sum(1 for r in lignes if r["tranche"] == t)}
                for t in ("<2 ans", "2-4 ans", "5-14 ans", "15-44 ans", "45-59 ans", "60 ans et plus")
                if any(r["tranche"] == t for r in lignes)],
        "sexe": g["sexe"],
        "sitrep_labo": periodes["sitrep"]["labo"],
    }
    with open("cholera_data.json", "w", encoding="utf-8") as f:
        json.dump(sortie, f, ensure_ascii=False)


def ecrire_series(lignes, periodes):
    """Letalite glissante sur trois semaines et pyramide des ages."""
    cles = periodes["cles"]["semaine"]
    sems = [int(c.split("-S")[1]) for c in cles]
    glissant = []
    for i, cle in enumerate(cles):
        fenetre = cles[max(0, i - 2):i + 1]
        n = sum(periodes["semaine"][c]["periode"]["cas_issue"] for c in fenetre)
        d = sum(periodes["semaine"][c]["periode"]["deces"] for c in fenetre)
        glissant.append({"sem": sems[i], "n": n, "d": d,
                         "cfr": round(d / n * 100, 1) if n >= 5 else None})
    cumule = [{"sem": int(c.split("-S")[1]), "cfr": periodes["semaine"][c]["cumul"]["letalite"]}
              for c in cles]
    ordre = ("<2 ans", "2-4 ans", "5-14 ans", "15-44 ans", "45-59 ans", "60 ans et plus")
    pyramide = []
    for t in ordre:
        h = sum(1 for r in lignes if r["tranche"] == t and r["sexe"] == "M")
        f = sum(1 for r in lignes if r["tranche"] == t and r["sexe"] == "F")
        if h + f:
            pyramide.append({"tranche": t, "H": h, "F": f})
    valides = [x for x in glissant if x["cfr"] is not None]
    actuel = valides[-1]["cfr"] if valides else None
    precedent = valides[-2]["cfr"] if len(valides) > 1 else None
    sortie = {
        "sems": sems, "cfr_glissant": glissant, "cfr_cumule": cumule,
        "cas_hebdo": [{"sem": int(c.split("-S")[1]),
                       "cas": periodes["semaine"][c]["periode"]["cas"],
                       "deces": periodes["semaine"][c]["periode"]["deces"]} for c in cles],
        "pyramide": pyramide,
        "variation": {
            "cfr_actuel": actuel, "cfr_precedent": precedent,
            "delta": round(actuel - precedent, 1) if actuel is not None and precedent is not None else None,
            "cas_actuel": periodes["semaine"][cles[-1]]["periode"]["cas"],
            "cas_precedent": periodes["semaine"][cles[-2]]["periode"]["cas"] if len(cles) > 1 else None,
            "sem_actuelle": sems[-1], "sem_precedente": sems[-2] if len(sems) > 1 else None,
            "cfr_cumule": periodes["global"]["letalite"],
        },
    }
    with open("chol_series.json", "w", encoding="utf-8") as f:
        json.dump(sortie, f, ensure_ascii=False)


def main():
    chemin = trouver_classeur()
    if not chemin:
        print("Aucun classeur cholera dans", BRUT, ": jeux de donnees inchanges.")
        return 0

    # Le referentiel du rapport de situation reste la source des confirmations
    # biologiques et des populations, le champ correspondant de la liste
    # lineaire presentant un codage heterogene.
    reference = {"numero": "026", "arrete": "2026-07-28", "cas": 218, "deces": 9,
                 "letalite": 4.1,
                 "labo": {"tdr_total": 42, "tdr_pos": 27,
                          "culture_total": 48, "culture_pos": 23},
                 "villages": 37, "zones_actives": 8, "taux_attaque": 14,
                 "population_totale": 1556246,
                 "populations": {"KARAL": 124300, "N'Djamena Nord": 165900}}
    if os.path.exists("chol_periodes.json"):
        with open("chol_periodes.json", encoding="utf-8") as f:
            reference = json.load(f).get("sitrep", reference)

    classeur = openpyxl.load_workbook(chemin, data_only=True)
    ws = feuille_liste_lineaire(classeur)
    if ws is None:
        print("Aucune feuille exploitable dans", chemin)
        return 0

    lignes = lire(ws)
    if not lignes:
        print("Liste lineaire vide dans", chemin)
        return 0
    normaliser_villages(lignes)

    periodes = ecrire_periodes(lignes, reference)
    ecrire_riposte(lignes, periodes)
    ecrire_series(lignes, periodes)
    g = periodes["global"]
    print(f"Cholera actualise depuis {os.path.basename(chemin)} : "
          f"{g['cas']} cas, {g['deces']} deces, letalite {g['letalite']}%, "
          f"{g['villages']} localites.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
